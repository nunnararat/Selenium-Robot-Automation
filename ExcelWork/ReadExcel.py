# Import Packate
import  openpyxl


# Load Workbook
wk = openpyxl.load_workbook("C:\\information\\TestingWorld\\Data\\TestData.xlsx")

print(wk.sheetnames)
print("Active Sheet is " + wk.active.title)


# Create Object of Any sheet on which you want to work
sh = wk['Sheet1']
print(sh.title)



print("---------- Read One Cell Data ----------")
print(sh['A3'].value)
print(sh['B4'].value)


# c1 = sh.cell(column=1, row=3)      #  column 1 , row 3 
c1 = sh.cell(3, 1)      # row3 , column 1
print(c1.value)

print("row of c1 : " , c1.row)
print("column of c1 : " , c1.column)



print("---------- Read All Row & Cell Data ----------")
# Find Rows having data
rows = sh.max_row
columns = sh.max_column

print("Total Rows are : ", str(rows))
print("Total Columns are : ", str(columns))

for row in range(1, rows+1):
    for column in range(1, columns+1):
        c = sh.cell(row, column)
        print(c.value)


for row in sh['A1':'C4']:
    for column in row:
        print(column.value)