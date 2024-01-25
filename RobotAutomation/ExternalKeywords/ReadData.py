import openpyxl

workbook = openpyxl.load_workbook("C://information//TestingWorld//Data//TestData.xlsx")


def fetch_number_of_rows(Sheetname):
    sheet = workbook[Sheetname]
    return sheet.max_row

def fetch_cell_data(Sheetname, row, cell):
    sheet = workbook[Sheetname]
    cell = sheet.cell(int(row), int(cell))
    return cell.value


# print(fetch_number_of_rows("Hello"))
# print(fetch_cell_data("Hello", 1, 1))               # Sheet => Hello , row 1, column 1






# sheet = workbook["Hello"]

# print(sheet.max_row)
# cell = sheet.cell(1,1)
# print(cell.value)